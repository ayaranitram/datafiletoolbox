# -*- coding: utf-8 -*-
"""
Created on Thu Jan 21 11:00:20 2021

@author: MCARAYA
"""

__version__ = '0.22.5'
__release__ = 220223
__all__ = ['XLSX']

from .mainObject import SimResult as _SimResult
from .._common.inout import _extension, _verbose
from .._common.functions import _mainKey
import pandas as pd
import os


class XLSX(_SimResult):
    """
    object to contain data read from .xlsx files

    """
    def __init__(self, inputFile=None, verbosity=2, sheet_name=None, header=[0, 1], units=1, overwrite=True, combine_SheetName_ColumnName=None, sheetName_auto='R', nameSeparator=':', ignore_nameSeparator=False, long=False, **kwargs) :
        _SimResult.__init__(self, verbosity=verbosity)
        self.kind = XLSX
        self.results = {}
        self.Frames = {}
        self.FramesIndex = {}
        self.overwrite = False
        self.lastFrame = ''
        if sheetName_auto.upper().strip() in ('L','R'):
            self.sheetName_auto = sheetName_auto.upper().strip()
        else:
            self.sheetName_auto = 'R'
        self.nameSeparator = str(nameSeparator)
        self.commonIndex = None
        self.commonIndexVector = None
        if 'frames' in kwargs or ( type(inputFile) is str and len(inputFile.strip()) > 0 ):
            if 'frames' in kwargs and kwargs['frames'] is not None:
                self.fromSimDataFrame(kwargs['frames'])
            elif os.path.isfile(inputFile) :
                self.readSimExcel(inputFile, sheet_name=sheet_name, header=header, units=units, combine_SheetName_ColumnName=combine_SheetName_ColumnName, ignore_nameSeparator=ignore_nameSeparator, long=long, **kwargs)
            else :
                print("file doesn't exists")
        if len(self.Frames) > 0 and inputFile is not None:
            self.name = _extension(inputFile)[1]
            # if 'TIME' not in self.keys :
            #     if 'DATE' in self.keys :
            #         TIME = np.array( [0] + list( self('DATE')[1:] - self('DATE')[:-1] ) )
            if type(long) is dict and 'index' in long and self.is_Key(long['index']) and 'index' not in kwargs:
                kwargs['index'] = long['index']
            self.initialize(**kwargs)

    def initialize(self, **kwargs) :
        """
        run intensive routines, to have the data loaded and ready
        """
        self.keys = tuple( sorted(self.keys) )
        self.extract_Wells()
        self.extract_Groups()
        self.extract_Regions()
        self.get_Attributes(None, True)
        self.find_index(**kwargs)
        self.otherDateNames()
        _SimResult.initialize(self, **kwargs)
        if self.is_Key('DATES') and not self.is_Key('DATE') :
            self['DATE'] = 'DATES'
        if not self.is_Key('DATE') :
            self.createDATES()
        elif self.get_Unit('DATE') is None or self.get_Unit('DATE') != 'DATE' :
            self.set_Units('DATE', 'DATE', overwrite=True)
        if not self.is_Key('DATES') and self.is_Key('DATE') :
            self['DATES'] = 'DATE'
        if self.is_Key('DATES') and ( self.get_Unit('DATES') is None or self.get_Unit('DATES') != 'DATE' ) :
            self.set_Unit('DATES', 'DATE', overwrite=True)
        if not self.is_Key('TIME') and self.is_Key('DATE') :
            self['TIME'] = ( self('DATE').astype('datetime64[s]') - self.start ).astype('int') / (60*60*24)
        if self.is_Key('TIME') and ( self.get_Unit('TIME') is None or self.get_Unit('TIME').upper() in ['', 'NONE'] ) :
            self.set_Unit('TIME', 'DAYS', overwrite=True)

    def find_index(self,**kwargs):
        """
        identify the column that is common to all the frames, to be used as index.
        If there is a single frame the first column is used.
        """

        possibleKeys = ('DATE', 'DATES', 'Date', 'date', 'Dates', 'dates', 'TIME', 'Time', 'time', 'DAYS', 'Days', 'days', 'MONTHS', 'Months', 'months', 'YEARS', 'Years', 'years') + self.keys  # + tuple(map(_mainKey,self.FramesIndex.keys()))
        if 'index' in kwargs and kwargs['index'] in self.FramesIndex:
            possibleKeys = (kwargs['index'],) + possibleKeys
            self.DTindex = kwargs['index']
        # possibleKeys = tuple(set(possibleKeys))

        # check current KeyIndex
        KeyIndex = True
        IndexVector = None
        _verbose(self.speak, 1, "looking for a common index column.")
        _verbose(self.speak, 1, " default index name is: " + str(self.DTindex))
        for frame in self.Frames:
            _verbose(self.speak, 1, "checking frame: " + str(frame))
            if self.DTindex not in self.FramesIndex:
                KeyIndex = False
                break
            elif self.FramesIndex[self.DTindex][1] not in self.Frames[frame] :
                KeyIndex = False
                break
            elif IndexVector is None :
                IndexVector = self.Frames[frame][self.FramesIndex[self.DTindex][1]]
            elif not IndexVector.equals( self.Frames[frame][self.FramesIndex[self.DTindex][1]] ):
                KeyIndex = False
                break
        if KeyIndex:
            _verbose(self.speak, 1, "found the common index: " + str(self.DTindex))
            return self.DTindex

        # look for other identical index
        for Key in possibleKeys:
            IndexVector = None
            if Key not in self.FramesIndex:  # and Key not in map(_mainKey,self.FramesIndex):
                KeyIndex = False
            else:
                KeyIndex = True
                for frame in self.Frames :
                    if self.FramesIndex[Key][1] not in self.Frames[frame].columns:
                        KeyIndex = False
                        break
                    elif IndexVector is None :
                        IndexVector = self.Frames[frame][self.FramesIndex[Key][1]]
                    elif not IndexVector.equals( self.Frames[frame][self.FramesIndex[Key][1]] ):
                        if 'date' in str(IndexVector.dtype) and IndexVector.dtype is self.Frames[frame][self.FramesIndex[Key][1]].dtype:
                            pass  # ii is a date, can be merged
                        else:
                            KeyIndex = False
                            break
            if KeyIndex :
                self.DTindex = Key
                break

        # look for other not indentical but common index
        if not KeyIndex:
            for Key in possibleKeys + self.keys :
                IndexVector = None
                if Key not in self.FramesIndex and Key not in map(_mainKey,self.FramesIndex):
                    KeyIndex = False
                    continue
                else:
                    KeyIndex = True
                    for frame in self.Frames :
                        if Key in self.FramesIndex :
                            if self.FramesIndex[Key][1] in self.Frames[frame].columns:
                                if IndexVector is None :
                                    IndexVector = self.Frames[frame][self.FramesIndex[Key][1]]
                                elif IndexVector.equals( self.Frames[frame][self.FramesIndex[Key][1]] ):
                                    pass
                                elif IndexVector.dtype is self.Frames[frame][self.FramesIndex[Key][1]].dtype and 'date' in str(IndexVector.dtype):
                                    pass
                                elif IndexVector.dtype is self.Frames[frame][self.FramesIndex[Key][1]].dtype and _mainKey(Key).upper().strip() == 'TIME':
                                    pass
                                else:
                                    KeyIndex = False
                                    break
                        elif _mainKey(Key) in self.FramesIndex:
                            if self.FramesIndex[_mainKey(Key)][1] in self.Frames[frame].columns:
                                if IndexVector is None :
                                    IndexVector = self.Frames[frame][self.FramesIndex[Key][1]]
                                elif IndexVector.equals( self.Frames[frame][self.FramesIndex[Key][1]] ):
                                    pass
                                elif str(IndexVector.dtype) == str(self.Frames[frame][self.FramesIndex[Key][1]].dtype) and 'date' in str(IndexVector.dtype):
                                    pass
                                elif str(IndexVector.dtype) == str(self.Frames[frame][self.FramesIndex[Key][1]].dtype) and _mainKey(Key).upper().strip() == 'TIME':
                                    pass
                                else:
                                    KeyIndex = False
                                    break
                if KeyIndex :
                    self.DTindex = Key
                    break

        if KeyIndex :
            _verbose(self.speak, 1, "found the common index: " + str(self.DTindex))
            # create result vector for the common index
            IndexVector = None
            for frame in self.Frames:
                if IndexVector is None:
                    IndexVector = self.Frames[frame][self.FramesIndex[self.DTindex][1]]
                elif IndexVector.equals(self.Frames[frame][self.FramesIndex[self.DTindex][1]]):
                    pass  # OK
                else:
                    IndexVector = pd.merge_ordered(IndexVector,self.Frames[frame][self.FramesIndex[self.DTindex][1]],how='outer').squeeze()
            self.add_Key(self.DTindex)
            self.TimeVector = self.DTindex
            _ = self.set_Vector(Key=self.DTindex, VectorData=IndexVector.to_numpy().reshape(-1,), Units=self.get_Units(self.DTindex), DataType='auto', overwrite=True)
            return self.DTindex
        else :
            _verbose(self.speak, 3, "not a common index name found.")
            self.DTindex = None

    def otherDateNames(self):
        """
        Looks for keys with 'datetime' or 'date' kind of data, merge all of them and creates a general DATE key.
        """
        if not self.is_Key('DATE') and not self.is_Key('DATES'):
            vector = None
            for datename in ('Date','date','Dates','dates','FECHA','Fecha','fecha','DATA','Data','data') + self.keys:
                if self.is_Key(datename):
                    if 'date' in str(self.get_RawVector(datename)[datename].dtype):
                        if vector is None:
                            vector = pd.Series(self.get_RawVector(datename)[datename],name='DATE')
                        else:
                            vector = pd.merge_ordered(vector, pd.Series(self.get_RawVector(datename)[datename],name='DATE'), how='outer').squeeze()
            _ = self.set_Vector(Key='DATE', VectorData=vector, Units='date', DataType='datetime', overwrite=True)

    def fromSimDataFrame(self,frame):
        self.units = frame.get_units()
        self.nameSeparator = frame.nameSeparator
        if frame.index.name is not None and frame.index.name not in frame.columns:
            indexName = frame.index.name
            self.DTindex = indexName
            frame.reset_index(inplace=True)
        self.Frames['sdf'] = frame
        self.name = 'from SimDataFrame'
        self.keys = frame.columns

    def readSimExcel(self, inputFile, sheet_name=None, header=[0, 1], units=1, combine_SheetName_ColumnName=None, ignore_nameSeparator=False, long=False, **kwargs) :
        """
        internal function to read an excel file with SimDataFrame format (header in first row, units in second row)
        """
        from .._dictionaries import ECL2VIPkey
        import numpy as np
        ECLkeys = tuple(ECL2VIPkey.keys())
        cleaningList = []
        # def _foundNewKey():
        #     self.FramesIndex[NewKey] = ( each, col )
        #     if col[-1].startswith('Unnamed:') :
        #         NewUnits = ''
        #         unitsMessage = ''
        #     else :
        #         NewUnits = col[-1].strip()
        #         unitsMessage = " with units: '"+NewUnits+"'"
        #     self.add_Key( NewKey )
        #     self.set_Unit( NewKey, NewUnits )
        #     _verbose(self.speak, 1, " > found key: '" + NewKey + "'" + unitsMessage )

        ### prepare header parameter
        if type(header) is int :
            if header < 0:
                raise ValueError("'header' parameter must be positive")
            header = [header]
        elif type(header) is not str and hasattr(header,'__iter__'):
            if not np.array([ h >= 0 for h in header]).all():
                raise ValueError("'header' parameter must be positive")
            header = list(header)
        else:
            raise TypeError("'header' parameter must be a positive integer or a list of positive integers.")
        ### prepare units parameter
        if type(units) is int:
            if units < 0:
                raise ValueError("'units' parameter must be positive")
            if units in header:
                if len(header) == 1:
                    _verbose(self.speak, 2, " > same row will be used as header and as units.")
                    header += [units]
                elif len(header) > 1:
                    header.pop(header.index(units))
                    header += [units]
            else:
                header += [units]
        elif units is None:
            header = header + [-1]
        else:
            raise TypeError("'units' parameter must be positive integer or None.")
        ### prepare a list of the labels rows parameter
        labels = [ r for r in header if r != units ]

        pdkwargs = kwargs.copy()
        for k in ['speak','verbosity','verbose','reload','preload','unload','ignoreSMSPEC','index','frames']:
            if k in pdkwargs:
                del pdkwargs[k]

        ### read the excel file using Pandas
        try:
            NewFrames = pd.read_excel(inputFile, sheet_name=sheet_name, header=header, **pdkwargs)
        except OSError:
            raise OSError("[Errno 24] Too many open files: " + str(inputFile) + '\nPlease close some other files.')
        except ImportError:
            raise ImportError("Missing optional dependencies 'xlrd' and 'openpyxl'.\nInstall xlrd and openpyxl for Excel support.\nUse pip or conda to install xlrd and install openpyxl.")
        except:
            try:
                import xlrd
                try:
                    import openpyxl
                except:
                    raise ModuleNotFoundError("Missing optional dependency 'openpyxl'.\nInstall openpyxl for Excel support.\nUse pip or conda to install openpyxl.")
            except:
                raise ModuleNotFoundError("Missing optional dependency 'xlrd'.\nInstall xlrd for Excel support.\nUse pip or conda to install xlrd.")
            raise TypeError('Not able to read the excel file, please check input parameters and excel sheets format.')

        ### check if only 1 sheet has been loaded, then put it into a dictionary
        if sheet_name is not None and type(NewFrames) is not dict:
            NewFrames = {str(sheet_name):NewFrames}

        ### if input file contains long format tables, pivot them to convert to wide tables
        if bool(long):
            from .._common.functions import _pivotDF
            if type(long) is dict:
                NewFrames = { sheet_name:_pivotDF(dataframe, **long) for sheet_name, dataframe in NewFrames.items() }
            elif type(long) is str:
                NewFrames = { sheet_name:_pivotDF(dataframe, long) for sheet_name, dataframe in NewFrames.items() }
            else:
                NewFrames = { sheet_name:_pivotDF(dataframe) for sheet_name, dataframe in NewFrames.items() }

        ### prepare combine_SheetName_ColumnName method
        ## the user has passed a parameter
        if bool(combine_SheetName_ColumnName):

            # a number must be the name of the sheet, put the number in a list
            if type(combine_SheetName_ColumnName) in (int, float):
                combine_SheetName_ColumnName = [combine_SheetName_ColumnName]

            # a keyword string
            elif type(combine_SheetName_ColumnName) is str and combine_SheetName_ColumnName.upper().strip() in ['RIGHT','LEFT','R','L','MAIN','ITEM']:
                # shout a message saying that the keyword could also be the name of a sheet
                if combine_SheetName_ColumnName in NewFrames:
                    _verbose(self.speak, 3, "WARNING: The paramater combine_SheetName_ColumnName is set to " + combine_SheetName_ColumnName +
                             " but this is also the name a sheet in the workbook.\n It will be considered as the " + combine_SheetName_ColumnName +
                             " option of the parameter. To set it as the sheet name provide the name of the sheet in a list: [" + combine_SheetName_ColumnName + "]")

                if combine_SheetName_ColumnName.upper().strip() in ['LEFT','L','MAIN']:
                    combine_SheetName_ColumnName = 'L'
                elif combine_SheetName_ColumnName.upper().strip() in ['RIGHT','R','ITEM']:
                    combine_SheetName_ColumnName = 'R'

            # a string that is the name of a sheet
            elif type(combine_SheetName_ColumnName) is str:
                if combine_SheetName_ColumnName not in NewFrames:
                    if combine_SheetName_ColumnName.strip() in NewFrames:
                        combine_SheetName_ColumnName = combine_SheetName_ColumnName.strip()
                    else:
                        raise ValueError("The sheetname '" + combine_SheetName_ColumnName + "' is not a sheet in the excel file \n " + str(inputFile))
                combine_SheetName_ColumnName = [combine_SheetName_ColumnName]

            # simply the boolean True, meaning to do it for every sheet in the workbook
            elif combine_SheetName_ColumnName is True:
                # shout a message saying there is only 1 sheet
                if len(NewFrames.keys()) == 1:
                    _verbose(self.speak, 2, " > You have requested to combine the name of the sheets with the names of the columns, but this excel file has only ONE sheet.\n If prefer not to combine the sheet name '" + str(list(NewFrames.keys())[0]) + "' with the name of every column set the parameter combine_SheetName_ColumnName=False")
                combine_SheetName_ColumnName = list(NewFrames.keys())

            # a list with the names of the sheets to be combined
            elif type(combine_SheetName_ColumnName) is list:
                for csn in range(len(combine_SheetName_ColumnName)):
                    # the list contains tuples of (sheetname, left or right) for each sheetname
                    if type(combine_SheetName_ColumnName[csn]) is tuple:
                        if combine_SheetName_ColumnName[csn][0] not in NewFrames:
                            if str(combine_SheetName_ColumnName[csn][0]) in NewFrames:
                                combine_SheetName_ColumnName[csn] = ( str(combine_SheetName_ColumnName[csn][0]) , str(combine_SheetName_ColumnName[csn][1]) )
                            elif str(combine_SheetName_ColumnName[csn][0]).strip() in NewFrames:
                                combine_SheetName_ColumnName[csn] = ( str(combine_SheetName_ColumnName[csn][0]).strip() , str(combine_SheetName_ColumnName[csn][1]) )
                            else:
                                raise ValueError("The sheetname '" + str(combine_SheetName_ColumnName[csn][0]) + "' is not a sheet in the excel file \n " + str(inputFile))
                    else:
                        if combine_SheetName_ColumnName[csn] not in NewFrames:
                            if str(combine_SheetName_ColumnName[csn]) in NewFrames:
                                combine_SheetName_ColumnName[csn] = str(combine_SheetName_ColumnName[csn])
                            elif str(combine_SheetName_ColumnName[csn]).strip() in NewFrames:
                                combine_SheetName_ColumnName[csn] = str(combine_SheetName_ColumnName[csn]).strip()
                            else:
                                raise ValueError("The sheetname '" + str(combine_SheetName_ColumnName[csn]) + "' is not a sheet in the excel file \n " + str(inputFile))

            # something different, try to cast into a list assuming there will be names of sheets...
            else:
                try:
                    combine_SheetName_ColumnName = list(combine_SheetName_ColumnName)
                except:
                    raise TypeError("not able to understand combine_SheetName_ColumnName parameter.")
                for csn in range(len(combine_SheetName_ColumnName)):
                    if combine_SheetName_ColumnName[csn] not in NewFrames:
                        if str(combine_SheetName_ColumnName[csn]) in NewFrames:
                            combine_SheetName_ColumnName[csn] = str(combine_SheetName_ColumnName[csn])
                        elif str(combine_SheetName_ColumnName[csn]).strip() in NewFrames:
                            combine_SheetName_ColumnName[csn] = str(combine_SheetName_ColumnName[csn]).strip()
                        else:
                            raise ValueError("The sheetname '" + str(combine_SheetName_ColumnName[csn]) + "' is not a sheet in the excel file \n " + str(inputFile))

        ## default automatic option
        elif combine_SheetName_ColumnName is None:

            # only 1 sheet, not combine
            if len(NewFrames.keys()) == 1:
                combine_SheetName_ColumnName = False

            # more than one sheet
            else:
                # list all the column names in the workook
                allColumns = []
                for nf in NewFrames:
                    allColumns += list(NewFrames[nf].columns)

                # if all the column names are different, no combination with sheetname is needed
                if len(set(allColumns)) == len(allColumns):
                    combine_SheetName_ColumnName = False

                # there are some repeated column names
                else:
                    # count the number of recognized keywords to try to guess if sheetname or column names are attributes
                    columnsKeys, sheetnameKeys = 0, 0
                    for k in ECLkeys + tuple([ I+str(ek)+H for I in 'FWGR' for ek in ECLkeys for H in ['','H']]):
                        sheetnameKeys += len( [ sn for sn in NewFrames.keys() if k in str(sn) ] )
                        columnsKeys += len( [ co for sn in NewFrames.keys() for co in NewFrames[sn] if k in str(co) ] )

                    if columnsKeys >= sheetnameKeys:  # sheetnames are item names
                        combine_SheetName_ColumnName = 'R'
                    else:
                        combine_SheetName_ColumnName = 'L'  # column names are item names

        ## any other option will be considered as False, no combination is mandated
        else:
            combine_SheetName_ColumnName = False

        ### process every frame read from the excel file
        for each in NewFrames:

            ## it's a new frame
            if each not in self.Frames:
                self.Frames[str(each)] = NewFrames[each]

                # define the name for every column
                for col in NewFrames[each].columns:
                    skip = False
                    NewKey = ' '.join(col[0:-1]).strip()  # all the column row but the units, units is put at the end... I think...
                    # NewKey = ' '.join([c for c in col if c in labels]).strip()

                    if bool(combine_SheetName_ColumnName) is True and self.nameSeparator in NewKey and not ignore_nameSeparator:
                        # nameSeparator already in the column name and ignore_nameSeparator=False
                        _verbose(self.speak, 2, " > identified 'MAIN:ITEM' structure in the column name '" + str(NewKey) + "' from sheet '" + str(each))

                    elif combine_SheetName_ColumnName == 'R':
                        # sheetname is ITEM name
                        _verbose(self.speak, 2, " > combined 'ColumnName:SheetName' to create MAIN:ITEM name structure for the column '" + str(NewKey) + "' from sheet '" + str(each))
                        NewKey = NewKey + self.nameSeparator + str(each).strip()

                    elif combine_SheetName_ColumnName == 'L':
                        # columnname is ITEM name
                        _verbose(self.speak, 2, " > combined 'SheetName:ColumnName' to create MAIN:ITEM name structure for sheet '" + str(each) + "' and the column '" + str(NewKey))
                        NewKey = str(each).strip() + self.nameSeparator + NewKey

                    elif type(combine_SheetName_ColumnName) is list:
                        # a list of sheetnames to be applied the combination of names
                        if ( each in combine_SheetName_ColumnName and self.sheetName_auto == 'R' ) \
                            or (each,'R') in combine_SheetName_ColumnName or (each,'r') in combine_SheetName_ColumnName:
                                _verbose(self.speak, 2, " > combined 'ColumnName:SheetName' to create MAIN:ITEM name structure for the column '" + str(NewKey) + "' from sheet '" + str(each))
                                NewKey = NewKey + self.nameSeparator + str(each).strip()

                        elif ( each in combine_SheetName_ColumnName and self.sheetName_auto == 'L' ) \
                            or (each,'L') in combine_SheetName_ColumnName or (each,'l') in combine_SheetName_ColumnName:
                                _verbose(self.speak, 2, " > combined 'SheetName:ColumnName' to create MAIN:ITEM name structure for sheet '" + str(each) + "' and the column '" + str(NewKey))
                                NewKey = str(each).strip() + self.nameSeparator + NewKey
                        else:
                            raise ValueError("Combination tuple of (sheetname, L or R) not understood for sheet " + str(each))

                    # defined the name of the column, write it in the index dictionary
                    if NewKey in self.FramesIndex:
                        # this column name already exist in the index

                        if NewFrames[each][col].equals( self.Frames[ self.FramesIndex[NewKey][0] ][ self.FramesIndex[NewKey][1] ] ):
                            # both columns contains the same data
                            skip = True
                            _verbose(self.speak, 2, " > skipping duplicated key: '" + NewKey + "' with duplicated data." )

                        elif NewKey.strip().upper() in ('TIME','DATE','YEAR','YEARS','MONTHS','DAYS'):
                            if min( self.Frames[ self.FramesIndex[NewKey][0] ][ self.FramesIndex[NewKey][1] ] ) <= min( NewFrames[each][col] ) and max( self.Frames[ self.FramesIndex[NewKey][0] ][ self.FramesIndex[NewKey][1] ] ) >= max( NewFrames[each][col] ):
                                skip = True
                                _verbose(self.speak, 1, " > skipping key: '" + NewKey + "' contained in data from previously loaded key." )

                        else:
                            # this new column contains new data
                            if bool(combine_SheetName_ColumnName) is False or self.nameSeparator in NewKey:
                                # not allowed to combine sheetname, iterate over a counter to append to the name
                                i = 1
                                while str(NewKey) + '_' + str(i).zfill(2) in self.FramesIndex:
                                    i += 1
                                _verbose(self.speak, 2, " > duplicated key: '" + NewKey + "' from sheet '" + str(each) + "' with new data was renamed to '" + str(NewKey) + '_' + str(i).zfill(2))
                                NewKey = str(NewKey) + '_' + str(i).zfill(2)

                            elif combine_SheetName_ColumnName == 'R':
                                i = None
                                if NewKey + self.nameSeparator + str(each).strip() in self.FrameIndex:
                                    i = 1
                                    while str(NewKey) + self.nameSeparator + str(each).strip() + '_' + str(i).zfill(2) in self.FramesIndex:
                                        i += 1
                                _verbose(self.speak, 2, " > duplicated key: '" + NewKey + "' from sheet '" + str(each) + "' with new data was renamed to '" + NewKey + self.nameSeparator + str(each).strip() + '_' + str(i).zfill(2))
                                NewKey = NewKey + self.nameSeparator + str(each).strip() + ('' if i is None else ('_' + str(i).zfill(2)))

                            elif combine_SheetName_ColumnName == 'L':
                                i = None
                                if str(each).strip() + self.nameSeparator + NewKey in self.FrameIndex:
                                    i = 1
                                    while str(each).strip() + self.nameSeparator + NewKey + '_' + str(i).zfill(2) in self.FramesIndex:
                                        i += 1
                                _verbose(self.speak, 2, " > duplicated key: '" + NewKey + "' from sheet '" + str(each) + "' with new data was renamed to '" + str(each).strip() + self.nameSeparator + NewKey + '_' + str(i).zfill(2))
                                NewKey = NewKey + self.nameSeparator + str(each).strip() + ('' if i is None else ('_' + str(i).zfill(2)))

                            elif type(combine_SheetName_ColumnName) is list:
                                # a list of sheetnames
                                if each in combine_SheetName_ColumnName:
                                    if self.sheetName_auto == 'R':
                                        i = None
                                        if NewKey + self.nameSeparator + str(each).strip() in self.FrameIndex:
                                            i = 1
                                            while str(NewKey) + self.nameSeparator + str(each).strip() + '_' + str(i).zfill(2) in self.FramesIndex:
                                                i += 1
                                        _verbose(self.speak, 2, " > duplicated key: '" + NewKey + "' from sheet '" + str(each) + "' with new data was renamed to '" + NewKey + self.nameSeparator + str(each).strip() + '_' + str(i).zfill(2))
                                        NewKey = NewKey + self.nameSeparator + str(each).strip() + ('' if i is None else ('_' + str(i).zfill(2)))

                                    elif self.sheetName_auto == 'L':
                                        i = None
                                        if str(each).strip() + self.nameSeparator + NewKey in self.FrameIndex:
                                            i = 1
                                            while str(each).strip() + self.nameSeparator + NewKey + '_' + str(i).zfill(2) in self.FramesIndex:
                                                i += 1
                                        _verbose(self.speak, 2, " > duplicated key: '" + NewKey + "' from sheet '" + str(each) + "' with new data was renamed to '" + str(each).strip() + self.nameSeparator + NewKey + '_' + str(i).zfill(2))
                                        NewKey = NewKey + self.nameSeparator + str(each).strip() + ('' if i is None else ('_' + str(i).zfill(2)))

                                # might be a list of tuples with left or right for each sheet
                                else:
                                    raise NotImplemented("automatic choose for list of tuples (sheetName, left or right) not yet implemented...")

                            else:
                                # at this point there should not be other options for combine_SheetName_ColumnName
                                raise TypeError("CODE ERROR: type of combine_SheetName_ColumnName variable not recognized: " + str(type(combine_SheetName_ColumnName)))

                            #     cleaningList.append(NewKey)  # to later remove the key

                            #     # rename the other Key
                            #     otherFrame = self.FramesIndex[NewKey][0]
                            #     if str(otherFrame) in ECLkeys or str(otherFrame)[1:4] in ECLkeys:
                            #         NewOtherKey = str(otherFrame).strip() + self.nameSeparator + NewKey
                            #     else:
                            #         NewOtherKey = NewKey + self.nameSeparator + str(otherFrame).strip()
                            #     if NewOtherKey not in self.FramesIndex:
                            #         self.FramesIndex[NewOtherKey] = ( self.FramesIndex[NewKey][0] , self.FramesIndex[NewKey][1] )
                            #         self.add_Key( NewOtherKey )
                            #         self.set_Unit( NewOtherKey, self.get_Unit(NewKey) )
                            #         _verbose(self.speak, 1, " > renamed key: '" + NewKey + "' to '" + NewOtherKey + "'")

                            #     # rename this frame key
                            #     if str(each) in ECLkeys or str(each)[1:4] in ECLkeys:
                            #         NewKey = str(each).strip() + self.nameSeparator + NewKey
                            #     else:
                            #         NewKey = NewKey + self.nameSeparator + str(each).strip()

                            # elif combine_SheetName_ColumnName is False or self.nameSeparator in str(otherFrame):
                            #     NewKey = NewKey + '_' + str( list(self.FramesIndex.keys()).count(NewKey) +1 )

                            # self.FramesIndex[NewKey] = ( each, col )
                            # if col[-1].startswith('Unnamed:') :
                            #     NewUnits = ''
                            #     unitsMessage = ''
                            # else :
                            #     NewUnits = col[-1].strip()
                            #     unitsMessage = " with units: '"+NewUnits+"'"
                            # self.add_Key( NewKey )
                            # self.set_Unit( NewKey, NewUnits )
                            # _verbose(self.speak, 1, " > found key: '" + NewKey + "'" + unitsMessage)

                    else:  # it is a new key, just put it in the index
                        pass  # nothing to do

                    if not skip:
                        # write the new key in the index
                        self.FramesIndex[NewKey] = ( each, col )
                        if col[-1].startswith('Unnamed:') :
                            NewUnits = ''
                            unitsMessage = ''
                        else :
                            NewUnits = col[-1].strip()
                            unitsMessage = " with units: '"+NewUnits+"'"
                        self.add_Key( NewKey )
                        self.set_Unit( NewKey, NewUnits )
                        _verbose(self.speak, 1, " > found key: '" + NewKey + "'" + unitsMessage)
                        # # _foundNewKey()
                        # if bool(combine_SheetName_ColumnName) is True and self.nameSeparator not in str(NewKey):
                        #     cleaningList.append(NewKey)  # to later remove the key

                        #     # rename the other Key
                        #     otherFrame = self.FramesIndex[NewKey][0]
                        #     if str(otherFrame) in ECLkeys or str(otherFrame)[1:4] in ECLkeys:
                        #         NewOtherKey = str(otherFrame).strip() + self.nameSeparator + NewKey
                        #     else:
                        #         NewOtherKey = NewKey + self.nameSeparator + str(otherFrame).strip()
                        #     if NewOtherKey not in self.FramesIndex:
                        #         self.FramesIndex[NewOtherKey] = ( self.FramesIndex[NewKey][0] , self.FramesIndex[NewKey][1] )
                        #         self.add_Key( NewOtherKey )
                        #         self.set_Unit( NewOtherKey, self.get_Unit(NewKey) )
                        #         _verbose(self.speak, 1, " > renamed key: '" + NewKey + "' to '" + NewOtherKey + "'")

            ## the frame name is already loaded, check if they are the same
            elif self.Frames[str(each)].equals(NewFrames[each]) :
                _verbose(self.speak, 2, "the sheet '"+each+"' was already loaded.")

            ## the frame name is already loaded but they are not the same
            else :
                ###############################################################
                ######### TO BE REVISITED #####################################
                ###############################################################
                if self.overwrite :
                    _verbose(self.speak, 2, "the sheet '"+str(each)+"' will overwrite the previously loaded sheet.")
                else :
                    i = 1
                    while str(each)+'_'+str(i).zfill(2) in self.Frames :
                        i += 1
                    _verbose(self.speak, 2, "the sheet '"+str(each)+"' will be loaded as '"+str(each)+'_'+str(i).zfill(2)+"' to not overwrite the previously loaded sheet.")
                    self.Frames[str(each)+'_'+str(i).zfill(2)] = NewFrames[each]
                    for col in NewFrames[each].columns :
                        NewKey = ' '.join(list(map(str,col[0:-1]))).strip()
                        self.FramesIndex[NewKey] = ( str(each)+'_'+str(i).zfill(2), col )
                        if col[-1].startswith('Unnamed:') :
                            NewUnits = ''
                            unitsMessage = ''
                        else :
                            NewUnits = col[-1].strip()
                            unitsMessage = " with units: '"+NewUnits+"'"
                        self.add_Key( NewKey )
                        userVerbose, self.speak = self.speak, 0
                        self.set_Unit( NewKey, NewUnits )
                        self.speak = userVerbose
                        _verbose(self.speak, 1, " > found key: '"+NewKey+"'" + unitsMessage )
                ###############################################################
                ######### END TO BE REVISITED #################################
                ###############################################################

        if len(cleaningList) > 0:
            self.keys = tuple(set( [ K for K in self.keys if K not in set(cleaningList) ] ))
            for K in cleaningList:
                if K in self.units:
                    del self.units[K]
                if K in self.FramesIndex:
                    del self.FramesIndex[K]

            columnsList = [ str(col[0] if type(col) is tuple else col) for frame in self.Frames for col in self.Frames[frame].columns  ]
            commonIndex = None
            for timeK in ['DATE','DATES','TIME','DAYS','MONTHS','YEARS']:
                if timeK in list(map(str.upper,map(_mainKey,columnsList))):
                    commonIndex = None
                    commonTimeK = timeK
                    for frame in self.Frames:
                        if timeK in [ str(col[0] if type(col) is tuple else col).strip().upper() for col in self.Frames[frame].columns ]:
                            candidate = self.Frames[frame][ [ col for col in self.Frames[frame].columns if str(col[0]).strip().upper() == timeK ][0] ]
                            candidate.rename(timeK, inplace=True)
                            if commonIndex is None:
                                commonIndex = candidate
                            else:
                                try:
                                    commonIndex = pd.merge_ordered( commonIndex, candidate, on=timeK, how='outer' )
                                except ValueError:
                                    if timeK in ['DATE','DATES']:
                                        raise ValueError('column ' + str(timeK) + ' has values that are not in date format in sheet ' + str(frame) + '.')
                                    else:
                                        raise ValueError('index column ' + str(timeK) + ' has different kind of values in sheet ' + str(frame) + '.')
                        else:
                            commonIndex = None
                            commonTimeK = None
                            break
                    if commonIndex is not None:
                        break
            if commonIndex is not None:
                if isinstance(commonIndex,pd.Series):
                    self.commonIndex = ( commonTimeK, pd.Index(commonIndex.values) )
                elif isinstance(commonIndex,pd.DataFrame):
                    self.commonIndex = ( commonTimeK, commonIndex.set_index(commonTimeK).index )
                self.FramesIndex[self.commonIndex[0]] = ( list(self.Frames.keys())[0], self.commonIndex[0] )
                for frame in self.Frames:
                    thisTimeK = [ col for col in self.Frames[frame].columns if str(col[0]).strip().upper() == self.commonIndex[0] ][0]
                    self.Frames[frame] = self.Frames[frame].set_index(thisTimeK, drop=True).reindex(index=self.commonIndex[1], method=None)
                    colsBefore = list(self.Frames[frame].columns)
                    self.Frames[frame] = self.Frames[frame].reset_index()
                    colsAfter = list(self.Frames[frame].columns)
                    indexName = [ c for c in colsAfter if c not in colsBefore ][0]
                    if type(indexName) is tuple:
                        indexName = indexName[0]
                    self.Frames[frame] = self.Frames[frame].rename(columns={indexName:self.commonIndex[0]})
                    if type(thisTimeK) is tuple:
                        partK = str(thisTimeK[0]).strip()
                    else:
                        partK = str(thisTimeK)

                    if partK + self.nameSeparator + frame in self.FramesIndex:
                        del self.FramesIndex[ partK + self.nameSeparator + frame ]
                    elif frame + self.nameSeparator + partK in self.FramesIndex:
                        del self.FramesIndex[ frame + self.nameSeparator + partK ]

                    if partK + self.nameSeparator + frame in self.units:
                        del self.units[ partK + self.nameSeparator + frame ]
                    elif frame + self.nameSeparator + partK in self.units:
                        del self.units[ frame + self.nameSeparator + partK ]

                    self.keys = tuple([ K for K in self.keys if K != (partK + self.nameSeparator + frame) and K != (frame + self.nameSeparator + partK) ])


    def list_Keys(self, pattern=None, reload=False) :
        """
        Return a StringList of summary keys matching @pattern.

        The matching algorithm is ultimately based on the fnmatch()
        function, i.e. normal shell-character syntax is used. With
        @pattern == "WWCT:*" you will get a list of watercut keys for
        all wells.

        If pattern is None you will get all the keys of summary
        object.
        """

        if pattern is None :
            return self.keys
        else:
            keysList = []
            for key in self.keys:
                if pattern in key :
                    keysList.append(key)
            return tuple( keysList )

    # support functions for get_Vector:
    def loadVector(self, key, Frame=None) :
        """
        internal function to return a numpy vector from the Frame files
        """
        if Frame is None and key not in self.FramesIndex:
            if ':' in key:
                if key.split(':')[0] in self.FramesIndex and key.split(':')[-1] in self.Frames:
                    return self.loadVector(key.split(':')[0], key.split(':')[-1])
                elif key.split(':')[-1] in self.FramesIndex and key.split(':')[0] in self.Frames:
                    return self.loadVector(key.split(':')[-1], key.split(':')[0])
        elif Frame is None and key in self.FramesIndex:
            Frame = self.FramesIndex[key][0]
        elif Frame in self.Frames:
            pass # OK
        else :
            _verbose(self.speak, 1, "the key '"+key+"' is not present in these frames.")
            return None
        if self.lastFrame == '':
            self.lastFrame = Frame

        thisFrame = self.lastFrame if Frame is None else Frame

        if key == self.DTindex or (self.FramesIndex[key][1] in self.Frames[thisFrame] and 'date' in str(self.Frames[thisFrame][self.FramesIndex[key][1]].dtype)):
            if thisFrame in self.Frames and key in self.Frames[thisFrame]:
                return self.Frames[thisFrame][self.FramesIndex[key][1]].to_numpy()
            elif key in self.FramesIndex:
                result = self.Frames[Frame][self.FramesIndex[key][1]]
                if len(result) == len(self.Frames[thisFrame]):
                    return result.to_numpy()
                else :
                    return self.Frames[thisFrame].index.to_numpy()
            else:
                return None
        elif key in self.FramesIndex:
            if len(self.Frames) == 1:
                result = self.Frames[thisFrame][self.FramesIndex[key][1]].to_numpy()
                self.lastFrame = thisFrame
            else:
                commonIndexVector = self.commonIndexVector
                if self.DTindex in self.Frames[thisFrame]:
                    result = self.Frames[thisFrame][[self.FramesIndex[key][1],self.DTindex]].set_index(self.DTindex)
                    if self.commonIndex is None or self.DTindex != self.commonIndex:
                        commonIndexVector = self.createCommonVector(self.DTindex)
                elif self.FramesIndex[self.DTindex][1] in self.Frames[thisFrame]:
                    result = self.Frames[thisFrame][[self.FramesIndex[key][1],self.FramesIndex[self.DTindex][1]]].set_index(self.FramesIndex[self.DTindex][1]).squeeze()
                    if self.commonIndex is None or self.DTindex != self.commonIndex:
                        commonIndexVector = self.createCommonVector(self.FramesIndex[self.DTindex][1])

                if commonIndexVector is not None:
                    self.commonIndex = self.DTindex
                    self.commonIndexVector = commonIndexVector
                    result = result.reindex(index=commonIndexVector)
                else:
                    result = self.Frames[thisFrame][self.FramesIndex[key][1]].to_numpy()
                self.lastFrame = Frame
            return result

    def createCommonVector(self, key):
        """
        look for the key on all the DataFrames and return a merged array
        """
        commonVector = None
        for frame in self.Frames:
            if key not in self.Frames[frame]:
                _verbose(self.speak, 1, "The key '" + str(key) + "' is not common to all the sheets.")
                return None
            if commonVector is None:
                commonVector = self.Frames[frame][key]
            else:
                commonVector = pd.merge_ordered(commonVector, self.Frames[frame][key], how='outer').squeeze()
        return commonVector.values

    # def extract_Wells(self) :
    #     """
    #     Will return a list of all the well names in case.

    #     If the pattern variable is different from None only groups
    #     matching the pattern will be returned; the matching is based
    #     on fnmatch(), i.e. shell style wildcards.
    #     """
    #     wellsList = [ K.split(':')[-1].strip() for K in self.keys if ( K[0] == 'W' and ':' in K ) ]
    #     wellsList = list( set( wellsList ) )
    #     wellsList.sort()
    #     self.wells = tuple( wellsList )

    #     return self.wells

    # def extract_Groups(self, pattern=None, reload=False) :
    #     """
    #     Will return a list of all the group names in case.

    #     If the pattern variable is different from None only groups
    #     matching the pattern will be returned; the matching is based
    #     on fnmatch(), i.e. shell style wildcards.
    #     """
    #     groupsList = [ K.split(':')[-1].strip() for K in self.keys if ( K[0] == 'G' and ':' in K ) ]
    #     groupsList = list( set( groupsList ) )
    #     groupsList.sort()
    #     self.groups = tuple( groupsList )
    #     if pattern is not None :
    #         results = []
    #         for group in self.groups :
    #             if pattern in group :
    #                 results.append(group)
    #         return tuple(results)
    #     else :
    #         return self.groups

    # def extract_Regions(self, pattern=None) :
    #     # preparing object attribute
    #     regionsList = [ K.split(':')[-1].strip() for K in self.keys if ( K[0] == 'G' and ':' in K ) ]
    #     regionsList = list( set( regionsList ) )
    #     regionsList.sort()
    #     self.groups = tuple( regionsList )
    #     if pattern is not None :
    #         results = []
    #         for group in self.groups :
    #             if pattern in group :
    #                 results.append(group)
    #         return tuple(results)
    #     else :
    #         return self.groups

    # def get_Unit(self, Key='--EveryType--') :
    #     """
    #     returns a string identifiying the unit of the requested Key

    #     Key could be a list containing Key strings, in this case a dictionary
    #     with the requested Keys and units will be returned.
    #     the Key '--EveryType--' will return a dictionary Keys and units
    #     for all the keys in the results file

    #     """
    #     if type(Key) is str and Key.strip() != '--EveryType--' :
    #         Key = Key.strip().upper()
    #         if Key in self.units :
    #             return self.units[Key]
    #         if Key in ['DATES','DATE'] :
    #                 self.units[Key] = 'DATE'
    #                 return 'DATE'
    #         # if Key in self.keys :
    #         #     return 'unitless'
    #         # else:
    #         #     UList = None
    #         #     if Key[0] == 'W' :
    #         #         UList=[]
    #         #         for W in self.get_Wells() :
    #         #             if Key+':'+W in self.units :
    #         #                 UList.append(self.units[Key+':'+W])
    #         #             elif Key in self.keys :
    #         #                 UList.append( self.results.unit(Key+':'+W) )
    #         #         if len(set(UList)) == 1 :
    #         #             self.units[Key] = UList[0]
    #         #             return UList[0]
    #         #         else :
    #         #             UList = None
    #         #     elif Key[0] == 'G' :
    #         #         UList=[]
    #         #         for G in self.get_Groups() :
    #         #             if Key+':'+G in self.units :
    #         #                 UList.append(self.units[Key+':'+G])
    #         #             elif Key in self.keys :
    #         #                 UList.append( self.results.unit(Key+':'+G) )
    #         #         if len(set(UList)) == 1 :
    #         #             self.units[Key] = UList[0]
    #         #             return UList[0]
    #         #         else :
    #         #             UList = None
    #         #     elif Key[0] == 'R' :
    #         #         UList=[]
    #         #         for R in self.get_Regions() :
    #         #             if Key+':'+R in self.units :
    #         #                 UList.append(self.units[Key+':'+R])
    #         #             elif Key in self.keys :
    #         #                 UList.append( self.results.unit(Key+':'+R) )
    #         #         if len(set(UList)) == 1 :
    #         #             self.units[Key] = UList[0]
    #         #             return UList[0]
    #         #         else :
    #         #             UList = None
    #         # if UList is None:
    #         #     if ':' in Key:
    #         #         if Key.split(':')[0] in self.FramesIndex and Key.split(':')[-1] in self.Frames:
    #         #             return self.get_Unit(Key.split(':')[0])
    #         #         elif Key.split(':')[-1] in self.FramesIndex and Key.split(':')[0] in self.Frames:
    #         #             return self.get_Unit(Key.split(':')[-1])
    #         if Key in self.keys :
    #             if ':' in Key :
    #                 if Key[0] == 'W' :
    #                     if Key.split(':')[-1] in self.wells :
    #                         return self.get_Unit(Key.split(':')[0])
    #                 if Key[0] == 'G' :
    #                     if Key.split(':')[-1] in self.groups :
    #                         return self.get_Unit(Key.split(':')[0])
    #                 if Key[0] == 'R' :
    #                     if Key.split(':')[-1] in self.regions :
    #                         return self.get_Unit(Key.split(':')[0])
    #             return None
    #         else:
    #             if Key[0] == 'W' :
    #                 UList=[]
    #                 for W in self.get_Wells() :
    #                     if Key+':'+W in self.units :
    #                         UList.append(self.units[Key+':'+W])
    #                 if len(set(UList)) == 1 :
    #                     self.units[Key] = UList[0]
    #                     return UList[0]
    #                 else :
    #                     return None
    #             elif Key[0] == 'G' :
    #                 UList=[]
    #                 for G in self.get_Groups() :
    #                     if Key+':'+G in self.units :
    #                         UList.append(self.units[Key+':'+G])
    #                 if len(set(UList)) == 1 :
    #                     self.units[Key] = UList[0]
    #                     return UList[0]
    #                 else :
    #                     return None
    #             elif Key[0] == 'R' :
    #                 UList=[]
    #                 for R in self.get_Regions() :
    #                     if Key+':'+R in self.units :
    #                         UList.append(self.units[Key+':'+R])
    #                 if len(set(UList)) == 1 :
    #                     self.units[Key] = UList[0]
    #                     return UList[0]
    #                 else :
    #                     return None
    #             UList = None

    #     elif type(Key) is str and Key.strip() == '--EveryType--' :
    #         Key = []
    #         KeyDict = {}
    #         for each in self.keys :
    #             if ':' in each :
    #                 Key.append( _mainKey(each) )
    #                 KeyDict[ _mainKey(each) ] = each
    #             else :
    #                 Key.append(each)
    #         Key = list( set (Key) )
    #         Key.sort()
    #         tempUnits = {}
    #         for each in Key :
    #             if each in self.units :
    #                 tempUnits[each] = self.units[each]
    #             elif each in self.keys and ( each != 'DATES' and each != 'DATE' ) :
    #                 if self.results.unit(each) is None :
    #                     tempUnits[each] = self.results.unit(each)
    #                 else :
    #                     tempUnits[each] = self.results.unit(each).strip('( )').strip("'").strip('"')
    #             elif each in self.keys and ( each == 'DATES' or each == 'DATE' ) :
    #                 tempUnits[each] = 'DATE'
    #             else :
    #                 if KeyDict[each] in self.units :
    #                     tempUnits[each] = self.units[KeyDict[each]]
    #                 elif KeyDict[each] in self.keys :
    #                     if self.results.unit(KeyDict[each]) is None :
    #                         tempUnits[each] = self.results.unit(KeyDict[each])
    #                     else :
    #                         tempUnits[each] = self.results.unit(KeyDict[each]).strip('( )').strip("'").strip('"')
    #         return tempUnits
    #     elif type(Key) in [list,tuple] :
    #         tempUnits = {}
    #         for each in Key :
    #             if type(each) == str :
    #                 tempUnits[each] = self.get_Unit(each)
    #         return tempUnits
